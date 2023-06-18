from playwright.sync_api import Playwright, sync_playwright, expect
import time
import openpyxl

# O código importa os módulos necessários: Playwright e sync_playwright do pacote playwright.sync_api, 
# expect do pacote playwright, time e openpyxl.

# Um arquivo do Excel chamado 'nome da planilha.xlsx' é carregado usando a função openpyxl.load_workbook. 
# Esse arquivo será o 'modelo' para a planilha que irá receber as informações da pagina.
# A planilha 'nome da planilha' dentro do arquivo é atribuída à variável processos_page.

book = openpyxl.load_workbook('nome do arquivo.xlsx')
book.create_sheet('nome da planilha')
processos_page = book['nome da planilha']

Dentro da função run, o navegador Chromium é inicializado usando playwright.chromium.launch(headless=False), 
o que significa que o navegador será visível durante a execução.

Um novo contexto é criado usando browser.new_context().

Uma nova página é aberta dentro do contexto usando context.new_page().

def run(playwright: Playwright) -> None:
    browser = playwright.chromium.launch(headless=False)
    context = browser.new_context()

    # Open new page
    page = context.new_page()
        # Go to https://...
    page.goto("AQUI SE COLOCA O LINK/URL DA PÁGINA")
    time.sleep(3) # COLOCA-SE UMA PAUSA NA EXECUÇÃO DO CÓDIGO PARA ESPERAR A PAGINA CARREGAR 
    # CASO TENHA QUE MARCAR ALGUMA OPÇÃO NA PAGINA PARA ESCOLHER AS INFORMAÇÕES QUE
    # QUEIRA OBTER.
    # FAZER A REPETIÇÃO DESSA AÇÃO DE ACORDO COM A QUANTIDADE DE INFORMAÇÕES QUE TENHA
    # QUE ESCOLHER
    page.locator("COLOCAR AQUI O XPATH DA INFORMAÇÃO").check()
    time.sleep(1.2)
    
    # CASO TENHA QUE COLOCAR ALGUM TERMO PARA FAZER A PROCURA DAS INFORMAÇÕES
    page.locator("COLOCAR AQUI O XPATH DA INFORMAÇÃ").fill("COLOCAR AS INFORMAÇÕES QUE SERÃO PROCURADAS")
    
    # CASO TENHA ALGUM BOTAO PARA SER CLICADO PRA PROCURAR AS INFORMAÇÕES
    # Click Avançar
    page.locator("COLOCAR AQUI O XPATH").click()    
    time.sleep(3)
   
    # QUANDO A FUNÇÃO LIC É CHAMADA, É INICIDO O LOOP..     
    z = 1
    while True:
        # A FUNÇÃO É RESPONSÁVEL POR EXTRAIR AS INFORMAÇÕES DE CADA PÁGINA.
        def lic():
            
            
            i = 0
            x = 1
            zz = 1
            
            while True:
                # Dentro do loop while True da função lic(), é verificado se x == "NUMERO". Se for verdadeiro, 
                # significa que todas as informações da página atual foram extraídas e é necessário avançar 
                # para a próxima página.
                if x == NUMERO:  # DEVE SER COLOCADO O NUMERO DE VEZES QUE UMA INFORMAÇAO É REPETIDA NA PAGINA.
                                 # CASO TENHA MAIS DE UMA INFORMACAO PARA SER OBTIDA, ELAS IRAO SE REPETIR O MESMO NÚMERO
                                 # DE VEZES, SENDO ESSE NUMERO QUE DEVE SER COLOCADO.
                   # Click text=Proxima
                    page.locator("COLOCAR AQUI O FULL XPATH DA BOTAO QUE IRÁ LEVAR PARA A PROXIMA PAGINA").click()                 
                    
                    xz = f'{zz}ª pagina'
                    processos_page.append([xz])
                    print(xz)
                    time.sleep(2)
                    zz += 1
                    x = 1
                                        
                # Equanto a condição x == x for verdadeira, significa que ainda há informações na página atual para extrair. 
                # O código usa seletores XPath para localizar os elementos desejados na página e extrair seu conteúdo 
                # usando text_content(). As informações são adicionadas à planilha 'nome da planilha' usando 
                # processos_page.append([aa, bb, cc]). Os contadores são atualizados para a próxima iteração.       
                elif x == x:             
                    aa = page.locator('Xpath').nth(i).text_content().strip()
                    bb = page.locator('Xpath').nth(i).text_content().strip()   
                    cc = page.locator(f'//html/body/table/tbody/tr[{n}]/td[1]/span[3]').nth(i).text_content().strip()                                   
                    processos_page.append([aa, bb, cc])                    
                    print(f'projeto nº: {aa}')                    
                    x += 1
                    book.save('nome da planilha.xlsx')        
        
        if z == 1:
            z+=1
            lic()               
        elif z == 15:
            break 
    
    # ---------------------
    context.close()
    browser.close()


with sync_playwright() as playwright:
    run(playwright)
